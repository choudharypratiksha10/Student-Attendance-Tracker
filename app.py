from flask import Flask, render_template, request, redirect, send_file
from openpyxl import Workbook, load_workbook
from datetime import datetime
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import os, io

app = Flask(__name__)

FILE = "attendance.xlsx"

# ---------------- CREATE FILE ----------------
if not os.path.exists(FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Roll", "Name", "Department", "Records"])
    wb.save(FILE)

# ---------------- HOME ----------------
@app.route('/')
def home():
    wb = load_workbook(FILE)
    ws = wb.active

    search = request.args.get("search", "").lower()
    dept_filter = request.args.get("department", "All")

    students = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        roll, name, dept, records = row

        if (search in str(roll).lower() or search in str(name).lower()) and \
           (dept_filter == "All" or dept == dept_filter):
            students.append((roll, name, dept))

    return render_template("index.html", students=students,
                           selected_dept=dept_filter, search=search)

# ---------------- ADD ----------------
@app.route('/add_student', methods=['POST'])
def add_student():
    name = request.form['name']
    roll = request.form['roll']
    dept = request.form['department']

    wb = load_workbook(FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == roll:
            return "Roll already exists!"

    ws.append([roll, name, dept, ""])
    wb.save(FILE)

    return redirect('/')

# ---------------- DELETE ----------------
@app.route('/delete/<roll>')
def delete_student(roll):
    wb = load_workbook(FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == roll:
            ws.delete_rows(row[0].row)
            break

    wb.save(FILE)
    return redirect('/')

# ---------------- EDIT ----------------
@app.route('/edit/<roll>')
def edit_student(roll):
    wb = load_workbook(FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == roll:
            return render_template("edit.html", student=row)

# ---------------- UPDATE ----------------
@app.route('/update', methods=['POST'])
def update_student():
    roll = request.form['roll']
    name = request.form['name']
    dept = request.form['department']

    wb = load_workbook(FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == roll:
            row[1].value = name
            row[2].value = dept
            break

    wb.save(FILE)
    return redirect('/')

# ---------------- MARK ATTENDANCE WITH DATE ----------------
@app.route('/mark_attendance', methods=['POST'])
def mark_attendance():
    roll = request.form['roll']
    status = request.form['status']
    date = request.form.get('date')

    if not date:
        from datetime import datetime
        date = datetime.now().strftime("%Y-%m-%d")

    wb = load_workbook(FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == roll:

            current = row[3].value or ""

            # prevent duplicate for same date
            if date in current:
                return "Attendance already marked for this date!"

            row[3].value = current + f"{date}:{status},"
            wb.save(FILE)

            return redirect('/')

    return "Student not found!"

# ---------------- REPORT ----------------
@app.route('/report')
def report():
    wb = load_workbook(FILE)
    ws = wb.active

    dept = request.args.get("department", "All")
    date = request.args.get("date")

    data = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        roll, name, d, records = row

        if dept != "All" and d != dept:
            continue

        records = records.split(",") if records else []
        records = [r for r in records if r]

        if date:
            records = [r for r in records if r.startswith(date)]

        total = len(records)
        present = sum("Present" in r for r in records)
        absent = sum("Absent" in r for r in records)
        percent = (present / total * 100) if total else 0

        data.append((roll, name, d, total, present, absent, f"{percent:.2f}%"))

    return render_template("report.html", data=data,
                           selected_dept=dept, selected_date=date)

# ---------------- PDF ----------------
@app.route('/download')
def download():
    wb = load_workbook(FILE)
    ws = wb.active

    dept = request.args.get("department", "All")
    date = request.args.get("date")

    data = [["Roll", "Name", "Dept", "Total", "Present", "Absent", "Percentage"]]

    for row in ws.iter_rows(min_row=2, values_only=True):
        roll, name, d, records = row

        if dept != "All" and d != dept:
            continue

        records = records.split(",") if records else []
        records = [r for r in records if r]

        if date:
            records = [r for r in records if r.startswith(date)]

        total = len(records)
        present = sum("Present" in r for r in records)
        absent = sum("Absent" in r for r in records)
        percent = (present / total * 100) if total else 0

        data.append([roll, name, d, total, present, absent, f"{percent:.2f}%"])

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer)
    styles = getSampleStyleSheet()

    elements = []
    title = "Attendance Report"
    if dept != "All":
        title += f" - {dept}"
    if date:
        title += f" ({date})"

    elements.append(Paragraph(title, styles['Title']))
    elements.append(Spacer(1, 20))

    table = Table(data)
    table.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('BACKGROUND', (0,0), (-1,0), colors.darkblue),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
    ]))

    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="report.pdf")

if __name__ == "__main__":
    app.run(debug=True)